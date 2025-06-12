from app.tool.base import BaseTool, ToolResult, ToolFailure
from app.exceptions import ToolError
from app.logger import logger
from typing import Dict, Any
import openpyxl

class ExcelEditorTool(BaseTool):
    """A tool for editing Excel files."""

    def __init__(self):
        super().__init__(
            name="excel_editor",
            description="A tool for reading and writing Excel files.",
 parameters={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "The path to the Excel file."
                    },
                    "operation": {
                        "type": "string",
                        "description": "The operation to perform (e.g., 'read_sheet', 'write_cell')."
                    },
                    # Add other parameters as needed for specific operations
                    "sheet_name": {
                        "type": "string",
                        "description": "The name of the sheet to operate on (for read_sheet, write_cell)."
                    },
                    "cell": {
                        "type": "string",
                        "description": "The cell address (e.g., 'A1') for write_cell."
                    },
 "value": {
                        "type": "any",
                        "description": "The value to write to the cell for write_cell."
                    },
 "sheet_data": {
 "type": "array",
 "description": "Data to write to a new sheet (list of lists)."
                    }
                },
                "required": ["file_path", "operation"]
            }
        )

    async def execute(self, file_path: str, operation: str, **kwargs: Any) -> ToolResult:
        """
        Executes the specified operation on the Excel file.

        Args:
 file_path: The path to the Excel file.

 Operations:
 - 'read_sheet': Reads data from a specified sheet.
 Parameters:
 - sheet_name (str, optional): The name of the sheet. Defaults to 'Sheet1'.
 - 'write_cell': Writes a value to a specific cell.
 Parameters:
 - sheet_name (str, optional): The name of the sheet. Defaults to 'Sheet1'.
 - cell (str): The cell address (e.g., 'A1').
 - value (any): The value to write.
 - 'create_sheet': Creates a new sheet and optionally writes data to it.
 Parameters:
 - sheet_name (str): The name of the new sheet.
 - sheet_data (list[list], optional): Data to populate the new sheet.
            operation: The operation to perform ('read_sheet', 'write_cell', etc.).
            **kwargs: Additional parameters for specific operations.

        Returns:
            A ToolResult object containing the result or a ToolFailure object in case of an error.
        """
        logger.info(f"Executing ExcelEditorTool operation: {operation} on file: {file_path}")

        try:
            workbook = None
            try:
                workbook = openpyxl.load_workbook(file_path)
            except FileNotFoundError:
                # For 'create_sheet' operation, it's okay if the file doesn't exist initially
                if operation != 'create_sheet':
                     error_message = f"Error: File not found at {file_path}"
                     logger.error(error_message)
                     return ToolFailure(error=error_message)
                else:
                    # Create a new workbook if the file doesn't exist for create_sheet
                    workbook = openpyxl.Workbook()
                    # Remove the default sheet created with a new workbook
                    if 'Sheet' in workbook.sheetnames:
                        default_sheet = workbook['Sheet']
                        workbook.remove(default_sheet)


            if operation == "read_sheet":
                sheet_name = kwargs.get("sheet_name", "Sheet1") # Default to 'Sheet1' if not provided
                if not sheet_name:
                    raise ToolError("sheet_name is required for 'read_sheet' operation.")

                if sheet_name not in workbook.sheetnames:
                    raise ToolError(f"Sheet '{sheet_name}' not found in '{file_path}'.")

                sheet = workbook[sheet_name]
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(list(row))
                return ToolResult(content=data)

            elif operation == "write_cell" or operation == "create_sheet":
                sheet_name = kwargs.get("sheet_name", "Sheet1") # Default to 'Sheet1' if not provided
                if not sheet_name:
                     raise ToolError("sheet_name is required for 'write_cell' and 'create_sheet' operations.")

                if operation == "write_cell":
                    cell = kwargs.get("cell")
                    value = kwargs.get("value")
                    if not all([cell, value is not None]):
                        raise ToolError("cell and value are required for 'write_cell' operation.")

                    if sheet_name not in workbook.sheetnames:
                         raise ToolError(f"Sheet '{sheet_name}' not found in '{file_path}'.")

                    sheet = workbook[sheet_name]
                    sheet[cell] = value
                    workbook.save(file_path)
                    return ToolResult(content=f"Successfully wrote value '{value}' to cell '{cell}' in sheet '{sheet_name}'.")

                elif operation == "create_sheet":
                    if sheet_name in workbook.sheetnames:
                         raise ToolError(f"Sheet '{sheet_name}' already exists in '{file_path}'.")

                    new_sheet = workbook.create_sheet(sheet_name)
                    sheet_data = kwargs.get("sheet_data")
                    if sheet_data:
                        for row_data in sheet_data:
                            new_sheet.append(row_data)

                    workbook.save(file_path)
                    return ToolResult(content=f"Successfully created sheet '{sheet_name}' in '{file_path}'.")

            else:
                raise ToolError(f"Unknown operation: {operation}")

        except ToolError as e:
            logger.error(f"Error in ExcelEditorTool: {e.message}")
             return ToolFailure(error=error_message)
        except Exception as e:
            logger.error(f"An unexpected error occurred in ExcelEditorTool: {e}")
            return ToolFailure(error=f"An unexpected error occurred: {e}")

# Example of how you might use it (for testing purposes, not part of the tool itself)
# async def main():
#     excel_tool = ExcelEditorTool()
#
#     # Example read operation
#     read_result = await excel_tool.execute(file_path="example.xlsx", operation="read_sheet", sheet_name="Sheet1")
#     print(f"Read result: {read_result.content if isinstance(read_result, ToolResult) else read_result.error}")
#
#     # Example write operation
#     write_result = await excel_tool.execute(file_path="example.xlsx", operation="write_cell", sheet_name="Sheet1", cell="A1", value="Hello MCP!")
#     print(f"Write result: {write_result.content if isinstance(write_result, ToolResult) else write_result.error}")
#
# if __name__ == "__main__":
#     import asyncio
#     asyncio.run(main())